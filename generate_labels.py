# -*- coding: utf-8 -*-
"""
生成弹幕标注Excel文件
从清洗后数据中随机抽样，生成供人工标注的xlsx文件
"""

import csv
import random
import os

random.seed(42)  # 固定随机种子，可复现

CLEAN_DIR = "danmaku_cleaned"
OUTPUT_FILE = "弹幕标注表.xlsx"

# 配置
SAMPLE_COUNT = {
    "科技": 135,
    "娱乐": 200,
    "汽车": 95,
}

# 读取所有清洗后数据
all_rows = []
for csv_file in sorted(os.listdir(CLEAN_DIR)):
    if not csv_file.endswith('.csv'):
        continue
    category = csv_file.split('_')[1].replace('.csv', '')
    path = os.path.join(CLEAN_DIR, csv_file)
    with open(path, 'r', encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
    
    count = SAMPLE_COUNT.get(category, 100)
    # 如果弹幕总数不够，就全部取
    if len(rows) <= count:
        sampled = rows
    else:
        sampled = random.sample(rows, count)
    
    for row in sampled:
        all_rows.append({
            '分区': category,
            'BV号': row['BV号'],
            '弹幕文本': row['弹幕文本'],
            '视频进度(秒)': row['视频进度(秒)'],
            '时间戳': row['时间戳'],
            'DMID': row['DMID'],
        })

print(f"总抽样: {len(all_rows)} 条")
print(f"  科技: {sum(1 for r in all_rows if r['分区'] == '科技')}")
print(f"  娱乐: {sum(1 for r in all_rows if r['分区'] == '娱乐')}")
print(f"  汽车: {sum(1 for r in all_rows if r['分区'] == '汽车')}")

# 写入Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "弹幕标注"
    
    # === 样式 ===
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='微软雅黑', size=10)
    text_align = Alignment(vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    
    # 正面标签颜色
    pos_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    neu_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    neg_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
    
    # === 表头 ===
    headers = ['序号', '分区', '弹幕文本', '视频进度', '标签']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # === 数据行 ===
    for i, row in enumerate(all_rows, 1):
        # 序号
        cell = ws.cell(row=i+1, column=1, value=i)
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # 分区
        cell = ws.cell(row=i+1, column=2, value=row['分区'])
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # 弹幕文本
        cell = ws.cell(row=i+1, column=3, value=row['弹幕文本'])
        cell.font = cell_font
        cell.alignment = text_align
        cell.border = thin_border
        
        # 视频进度
        cell = ws.cell(row=i+1, column=4, value=f"{float(row['视频进度(秒)']):.1f}s")
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # 标签（空，等人工填）
        cell = ws.cell(row=i+1, column=5, value='')
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # === 数据验证（下拉菜单）===
    # E2到最后一行，只能选 正面/中性/负面
    dv = DataValidation(type="list", formula1='"正面,中性,负面"', allow_blank=True)
    dv.error = "请选择：正面、中性、负面"
    dv.errorTitle = "无效标签"
    dv.prompt = "请选择标签"
    dv.promptTitle = "情感标签"
    ws.add_data_validation(dv)
    dv.add(f'E2:E{len(all_rows)+1}')
    
    # === 列宽 ===
    ws.column_dimensions['A'].width = 6     # 序号
    ws.column_dimensions['B'].width = 8     # 分区
    ws.column_dimensions['C'].width = 60    # 弹幕文本
    ws.column_dimensions['D'].width = 10    # 视频进度
    ws.column_dimensions['E'].width = 10    # 标签
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # === 标注说明sheet ===
    ws2 = wb.create_sheet("标注说明")
    instructions = [
        ["弹幕情感标注说明"],
        [""],
        ["标注任务："],
        ["请为每条弹幕选择一个情感标签：正面、中性、负面"],
        [""],
        ["标签定义："],
        ["正面 — 表达赞赏、喜爱、支持、认同、期待、幽默、调侃（不含恶意）等积极情感"],
        ["中性 — 客观描述、提问、感叹、无明确情感倾向的陈述"],
        ["负面 — 表达不满、批评、嘲讽（恶意）、反对、质疑、焦虑、失望等消极情感"],
        [""],
        ["注意事项："],
        ["1. 括号标签列（E列）有下拉菜单，直接选择即可"],
        ["2. 如果一条弹幕同时包含正面和负面，看主要情感倾向"],
        ["3. \"哈哈哈哈\"类的幽默弹幕归为正面（表达开心）"],
        ["4. \"666\"\"厉害\"等归为正面（表达赞赏）"],
        ["5. \"？\"\"无语\"等归为负面（表达困惑/不满）"],
        ["6. 纯信息性内容（如\"今天几号\"）归为中性"],
        ["7. 讽刺/阴阳怪气归为负面"],
        [""],
        ["数据统计："],
        [f"总标注量: {len(all_rows)} 条"],
        [f"科技: {sum(1 for r in all_rows if r['分区'] == '科技')} 条"],
        [f"娱乐: {sum(1 for r in all_rows if r['分区'] == '娱乐')} 条"],
        [f"汽车: {sum(1 for r in all_rows if r['分区'] == '汽车')} 条"],
        [""],
        ["预计用时: 20-30分钟"],
    ]
    
    for row_idx, (text,) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(name='微软雅黑', size=14, bold=True)
        elif text.startswith("标签定义") or text.startswith("注意事项") or text.startswith("数据统计"):
            cell.font = Font(name='微软雅黑', size=11, bold=True)
        else:
            cell.font = Font(name='微软雅黑', size=10)
    
    ws2.column_dimensions['A'].width = 70
    
    wb.save(OUTPUT_FILE)
    print(f"\n标注文件已生成: {OUTPUT_FILE}")
    print("请用Excel打开，在E列下拉选择标签即可")

except ImportError:
    print("openpyxl未安装，正在安装...")
    import subprocess
    subprocess.check_call([
        r'C:\Users\SOLO\.workbuddy\binaries\python\envs\default\Scripts\pip.exe',
        'install', 'openpyxl', '-q'
    ])
    print("安装完成，请重新运行脚本")
